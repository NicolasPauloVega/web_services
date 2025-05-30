# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.chart import BarChart, Reference
# from io import BytesIO
# from .models import Usuarios

# def generar_excel_usuarios():
#     # Crear libro y hoja
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Estadísticas Usuarios"

#     # Encabezado de la tabla
#     encabezados = ["Nombre", "Puntos"]
#     ws.append(encabezados)

#     # Aplicar estilos al encabezado
#     for col in range(1, len(encabezados) + 1):
#         cell = ws.cell(row=1, column=col)
#         cell.font = Font(bold=True, color="FFFFFF")
#         cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
#         cell.alignment = Alignment(horizontal="center")

#     # Obtener datos de usuarios
#     usuarios = Usuarios.objects.values("nombres", "puntos_totales")

#     if not usuarios.exists():
#         return None  # No hay datos para exportar

#     max_puntos = max(u["puntos_totales"] for u in usuarios)
#     min_puntos = min(u["puntos_totales"] for u in usuarios)

#     # Llenar la tabla
#     for index, usuario in enumerate(usuarios, start=2):
#         ws.append([usuario["nombres"], usuario["puntos_totales"]])

#         # Resaltar en verde el máximo y en rojo el mínimo
#         puntos_cell = ws.cell(row=index, column=2)
#         if usuario["puntos_totales"] == max_puntos:
#             puntos_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
#         elif usuario["puntos_totales"] == min_puntos:
#             puntos_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

#     # Crear gráfico de barras con los puntos de todos los usuarios
#     chart = BarChart()
#     chart.title = "Puntos por Usuario"
#     chart.x_axis.title = "Usuarios"
#     chart.y_axis.title = "Puntos"

#     nombres = Reference(ws, min_col=1, min_row=2, max_row=len(usuarios) + 1)
#     puntos = Reference(ws, min_col=2, min_row=1, max_row=len(usuarios) + 1)

#     chart.add_data(puntos, titles_from_data=True)
#     chart.set_categories(nombres)
#     ws.add_chart(chart, "D2")  # Ubicar el gráfico en la celda D2

#     # Guardar en memoria
#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)

#     return output

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from .models import Usuarios, Programa, Puntos, Actividad

def generar_excel_usuarios():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estadísticas Usuarios"

    fila_actual = 1  # Mantener la fila para insertar

    # Recorremos cada programa
    for programa in Programa.objects.all():
        actividades = list(programa.actividades.all())
        actividades_nombres = [a.nombre for a in actividades]

        # Encabezados
        encabezados = ["Usuario", "Programa", "Puntaje Total"] + actividades_nombres
        ws.append(encabezados)

        # Estilos al encabezado
        for col in range(1, len(encabezados) + 1):
            cell = ws.cell(row=fila_actual, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        fila_actual += 1

        # Usuarios que tienen puntos en este programa
        puntos_programa = Puntos.objects.filter(programa=programa).select_related('usuario', 'actividad')
        usuarios = set(p.usuario for p in puntos_programa)

        for usuario in usuarios:
            fila = [f"{usuario.nombres} {usuario.apellidos}", programa.nombre, usuario.puntos_totales]

            # Inicializar dict de actividades con valor vacío
            actividad_puntajes = {a.nombre: '' for a in actividades}

            # Obtener puntajes del usuario en este programa
            puntos_usuario = puntos_programa.filter(usuario=usuario)

            for punto in puntos_usuario:
                actividad_puntajes[punto.actividad.nombre] = punto.puntos

            fila += [actividad_puntajes[nombre] for nombre in actividades_nombres]
            ws.append(fila)
            fila_actual += 1

        fila_actual += 2  # Espacio entre programas

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
